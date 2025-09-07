from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import time
import os
import re
import shutil

print("Abriendo Vacantes:")
# Cargar el archivo Excel
archivo_excel = "vacantes.xlsx"
workbook = load_workbook(archivo_excel)

# Seleccionar la hoja activa (o una específica)
hoja = workbook.active  # También puedes usar workbook["NombreHoja"]
ministerio = ""

# Crear un nuevo libro y hoja para la salida
wb_salida = Workbook()
hoja_salida = wb_salida.active
hoja_salida.title = "Convertido by Josefer"

# Escribir cabeceras en la primera fila
cabeceras = [
    "Puesto Numero",
    "Ministerio",
    "Centro Directivo",
    "Provincia",
    "localidad",
    "Puesto de Trabajo",
    "RPT",
    "Nivel",
    "Complemento"
]
hoja_salida.append(cabeceras)
# Crear autofiltros en la primera fila
hoja_salida.auto_filter.ref = hoja_salida.dimensions

fila_salida = 2  # Contador de filas para la hoja de salida
print("Procesando Filas:")
# Recorrer fila por fila y columna por columna
for fila in hoja.iter_rows(values_only=True):  # values_only=True devuelve solo los valores
    if fila[0] is None or str(fila[0]).strip() == "":
        continue
    if str(fila[0]).strip().upper().startswith("PUESTO"):
        continue  # Si la primera columna comienza con "PUESTO", saltar a la siguiente fila
    if isinstance(fila[0], str) and not fila[0].strip().isdigit():
        ministerio = fila[0].strip()
        print(ministerio)
        continue
    columna_salida = 1  # Contador de columnas para la hoja de salida
    for celda in fila:
        if celda is None or (isinstance(celda, str) and celda.strip() == ""):
            continue  # Si la columna está en blanco, saltar la columna
        hoja_salida.cell(row=fila_salida, column=columna_salida, value=celda)
        hoja_salida.cell(row=fila_salida, column=columna_salida).alignment = Alignment(wrap_text=True)
# Si es la columna del numero del puesto, guardar el ministerio en la segunda columna
        if columna_salida == 1:
            hoja_salida.cell(row=fila_salida, column=2, value=ministerio)  # Guardar ministerio en la seguna columna
            columna_salida += 1
# Si es la columna de la Provincia y localidad, guardar la localidad en la siguente columna
        if columna_salida == 4:
            partes = celda.split('\n', 1)
            hoja_salida.cell(row=fila_salida, column=4, value=partes[0])
            columna_salida += 1
            hoja_salida.cell(row=fila_salida, column=5, value=partes[1])
# Si es la columna de la puesto de trabajo extraigo el rpt
        if columna_salida == 6:
            # Extraer los últimos dígitos numéricos de la columna 6
            texto = celda if isinstance(celda, str) else ""
            ultimos_numeros = ""
            numeros = re.findall(r'\d+', texto)
            if numeros:
                ultimos_numeros = numeros[-1]
            columna_salida += 1
            hoja_salida.cell(row=fila_salida, column=7, value=int(ultimos_numeros))
# Si es la columna de nivel y complemento la reparto
        if columna_salida == 8:
            partes = celda.split('\n', 1)
            hoja_salida.cell(row=fila_salida, column=8, value=int(partes[0]))
            columna_salida += 1
            valor_complemento = partes[1].replace('.', '').replace(',', '.')
            try:
                valor_complemento = float(valor_complemento)
            except ValueError:
                valor_complemento = 0.0
            hoja_salida.cell(row=fila_salida, column=9, value=valor_complemento)
        columna_salida += 1
    fila_salida += 1

       
# Aplicar formato numérico a las columnas G, H (entero) e I (2 decimales)
for fila in hoja_salida.iter_rows(min_row=2, min_col=7, max_col=9):
    if fila[0]:  # Columna G
        fila[0].number_format = '0'
    if fila[1]:  # Columna H
        fila[1].number_format = '0'
    if fila[2]:  # Columna I
        fila[2].number_format = '0.00'
        
# Ajustar el ancho de las columnas al texto que contienen
for columna in hoja_salida.columns:
    max_length = 0
    columna_letra = columna[0].column_letter  # Obtener la letra de la columna
    for celda in columna:
        if celda.value:
            # Convertir a string y considerar saltos de línea
            for linea in str(celda.value).split('\n'):
                if len(linea) > max_length:
                    max_length = len(linea)
    # Ajustar el ancho (puedes sumar un extra para margen visual)
    hoja_salida.column_dimensions[columna_letra].width = max_length + 2

# Ajustar el alto de las filas al texto que contienen (considerando saltos de línea)
for fila in hoja_salida.iter_rows():
    max_lineas = 1
    for celda in fila:
        if celda.value and isinstance(celda.value, str):
            lineas = celda.value.count('\n') + 1
            if lineas > max_lineas:
                max_lineas = lineas
    hoja_salida.row_dimensions[fila[0].row].height = max_lineas * 15  # 15 es un valor estándar por línea

# Nombre dl fichero de salida
salida = "salida_filtrada.xlsx"

# Esperar hasta que el fichero de salida esté libre
while os.path.exists(salida):
    try:
        os.rename(salida, salida)
        break  # Si no hay excepción, el archivo está libre
    except OSError:
        print(f"El archivo '{salida}' está en uso. Ciérralo para continuar...")
        time.sleep(3)  # Espera 3 segundos antes de volver a intentar
            
wb_salida.save("salida_filtrada.xlsx")
# Comprobar que el fichero de destino no está en uso antes de copiarlo
destino = "salida_sin_rpt_info.xlsx"
while os.path.exists(destino):
    try:
        os.rename(destino, destino)
        break  # Si no hay excepción, el archivo está libre
    except OSError:
        print(f"El archivo '{destino}' está en uso. Ciérralo para continuar...")
        time.sleep(3)  # Espera 3 segundos antes de volver a intentar

shutil.copyfile("salida_filtrada.xlsx", destino)
print("Archivo salida_sin_rpt_info creado correctamente.")

print("Añadiendo información de las rpt:")
# Abrir el archivo rpt.xlsx
rpt_wb = load_workbook("rpt.xlsx", data_only=True)
rpt_hoja = rpt_wb.active

# Crear un diccionario para buscar rápido por columna 13 (M)
rpt_dict = {}
for row in rpt_hoja.iter_rows(min_row=2, values_only=True):  # min_row=2 si hay cabecera
    clave = row[12]  # Columna 13 (índice 12)
    if clave is not None:
        rpt_dict[str(clave).strip()] = row[20:32]  # Columnas U (21) a AF (32), índices 20 a 31

# Ahora abre tu archivo de salida para añadir los datos
wb_salida = load_workbook("salida_filtrada.xlsx")
hoja_salida = wb_salida.active

# Leer cabeceras reales de rpt.xlsx para columnas U (21) a AF (32)
rpt_cabeceras = []
for cell in rpt_hoja[1][20:32]:  # Índices 20 a 31 (columnas U a AF)
    rpt_cabeceras.append(cell.value if cell.value else f"RPT_{cell.column_letter}")

# Añadir las cabeceras reales a la hoja de salida
for idx, cab in enumerate(rpt_cabeceras, start=10):  # Empieza en la columna 10 (J)
    hoja_salida.cell(row=1, column=idx, value=cab)

hoja_salida.cell(row=1, column=22, value="Descripcion Observacion 1")
hoja_salida.cell(row=1, column=23, value="Descripcion Observacion 2")
hoja_salida.cell(row=1, column=24, value="Descripcion Observacion 3")
hoja_salida.cell(row=1, column=25, value="Descripcion Observacion 4")
hoja_salida.cell(row=1, column=26, value="Descripcion Observacion 5")

# Recorre las filas de salida y busca el valor en rpt.xlsx
for row in hoja_salida.iter_rows(min_row=2):
    valor_col7 = row[6].value  # Columna 7 (índice 6)
    if valor_col7 is None:
        continue
    datos_rpt = rpt_dict.get(str(valor_col7).strip())
    if datos_rpt:
        for i, dato in enumerate(datos_rpt):
            row[9 + i].value = dato  # Empieza en la columna 10 (índice 9)
                # Procesar columna T (índice 19, columna 20) - Observaciones

# Ahora, también procesamos la columna T (índice 19, columna 20) de salida_filtrada.xlsx
for row in hoja_salida.iter_rows(min_row=2):
    observaciones = row[19].value  # Columna T (índice 19)
    if observaciones and isinstance(observaciones, str):
        partes_obs = [parte.strip() for parte in observaciones.split(',')]
        for j, parte in enumerate(partes_obs):
            # Columna V es la 22, índice 21
            row[21 + j].value = parte

        
# Cargar claves.xlsx y crear diccionario de códigos
claves_wb = load_workbook("claves.xlsx", data_only=True)
claves_hoja = claves_wb.active
claves_dict = {}
for row in claves_hoja.iter_rows(min_row=2, values_only=True):  # min_row=2 para saltar cabecera
    codigo = str(row[0]).strip() if row[0] is not None else ""
    texto = str(row[1]).strip() if row[1] is not None else ""
    if codigo:
        claves_dict[codigo] = texto

# Procesar columnas V (índice 21) a Z (índice 25) de salida_filtrada.xlsx
for row in hoja_salida.iter_rows(min_row=2):
    for col in range(21, 26):
        valor = row[col].value
        if valor and isinstance(valor, str):
            codigo = valor.strip().split()[0].replace(",", "")  # Extrae el código (antes del primer espacio o coma)
            texto = claves_dict.get(codigo, valor)
            row[col].value = texto

for col in ['V', 'W', 'X', 'Y', 'Z']:
    hoja_salida.column_dimensions[col].width = 40
for col in ['T']:
    hoja_salida.column_dimensions[col].width = 12

wb_salida.save("salida_filtrada.xlsx")

# Comprobar que el fichero de destino no está en uso antes de copiarlo
destino = "salida_con_rpt_info.xlsx"
while os.path.exists(destino):
    try:
        os.rename(destino, destino)
        break  # Si no hay excepción, el archivo está libre
    except OSError:
        print(f"El archivo '{destino}' está en uso. Ciérralo para continuar...")
        time.sleep(3)  # Espera 3 segundos antes de volver a intentar

shutil.copyfile("salida_filtrada.xlsx", destino)
print("Fichero salida_con_rpt_info.xlsx generado.")
os.remove("salida_filtrada.xlsx")

import subprocess
subprocess.call(["python", "david.py"])