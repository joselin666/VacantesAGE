from openpyxl import load_workbook
from openpyxl.styles import Alignment
import time
import os

# Abrir el archivo de entrada
wb = load_workbook("salida_con_rpt_info.xlsx")
ws = wb.active

# 1. Eliminar columnas J (10) hasta U (21) inclusive
for _ in range(10, 22):
    ws.delete_cols(10)

# 2. Fusionar columnas J (10) hasta N (14) en la columna J (10)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    valores = []
    for col in range(10, 15):  # J=10, K=11, L=12, M=13, N=14
        valor = row[col - 1].value
        if valor:
            valores.append(str(valor))
    row[9].value = "\n".join(valores) if valores else ""
    # Limpiar las columnas K-N
    for col in range(11, 15):
        row[col - 1].value = None

# 3. Eliminar columnas K (11) hasta N (14)
for _ in range(4):
    ws.delete_cols(11)

# 4. Insertar la columna "mi orden" al principio
ws.insert_cols(1)
ws.cell(row=1, column=1, value="mi orden")
# Elimina la primera fila (la fila 1, usualmente títulos)
ws.delete_rows(1)
    
# 5. Guardar el archivo como salida_david.xlsx
# Ajustar el ancho de las columnas al texto que contienen
for columna in ws.columns:
    max_length = 0
    columna_letra = columna[0].column_letter  # Obtener la letra de la columna
    for celda in columna:
        if celda.value:
            # Convertir a string y considerar saltos de línea
            for linea in str(celda.value).split('\n'):
                if len(linea) > max_length:
                    max_length = len(linea)
    # Ajustar el ancho (puedes sumar un extra para margen visual)
    ws.column_dimensions[columna_letra].width = max_length + 2

# Ajustar el alto de las filas al texto que contienen (considerando saltos de línea)
for fila in ws.iter_rows():
    max_lineas = 1
    for celda in fila:
        if celda.value and isinstance(celda.value, str):
            lineas = celda.value.count('\n') + 1
            if lineas > max_lineas:
                max_lineas = lineas
    ws.row_dimensions[fila[0].row].height = max_lineas * 15  # 15 es un valor estándar por línea
# Limitar el ancho de la columna k a 60
ws.column_dimensions['K'].width = 60
            
# Comprobar que el fichero de destino no está en uso antes de copiarlo
destino = "salida_david.xlsx"
while os.path.exists(destino):
    try:
        os.rename(destino, destino)
        break  # Si no hay excepción, el archivo está libre
    except OSError:
        print(f"El archivo '{destino}' está en uso. Ciérralo para continuar...")
        time.sleep(3)  # Espera 3 segundos antes de volver a intentar

wb.save("salida_david.xlsx")
print("Archivo salida_david.xlsx generado correctamente.")